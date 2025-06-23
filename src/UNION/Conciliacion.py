import numpy as np
import pandas as pd
from statsmodels.tsa.filters.hp_filter import hpfilter
from openpyxl import load_workbook
import win32com.client
import time

# Path to Excel file
ruta_excel = r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx"

# Launch Excel and open the file
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

# Open and update all links (UpdateLinks=3 updates everything)
wb = excel.Workbooks.Open(Filename=ruta_excel, UpdateLinks=3)

# Optional: Wait to ensure updates complete (useful for async links)
time.sleep(2)

# Save with updated values
wb.Save()
wb.Close()
excel.Quit()

# Definición de parámetros
periodos = {
    "301": "2007.1",
    "302": "2024.3",
    "303": "2025.4",
    "201": "2007-01-01",
    "202": "2024-09-01",
    "203": "2025-12-01",
}

# Definición de valores enteros
anos_c = 17  # Años completos desde 2007 hasta el último año completo de la CTR
an = 19  # Años disponibles de datos mensualizados desde 2007 hasta 2025
tri = an * 4
mes = an * 12
ser = 10
p_ofe = 0.5
p_dem = 1 - p_ofe

# Creación de estructuras de datos
V4 = np.ones((4, 1))
v12 = np.ones((12, 1))

HTW = np.zeros((tri, tri * ser))  # Matriz HTW
HMW = np.zeros((tri * (ser - 1) + mes, mes * ser))  # Matriz HMW

CERO = np.zeros((tri, 1))
agre = np.full((1, 3), 1/3)
Bagre = np.kron(np.eye(tri), agre)
Bmagre = np.kron(np.eye(tri * (ser - 1)), agre)

Difer = np.diag(np.ones(mes)) + np.diag(-np.ones(mes - 1), -1)
dmas = np.ones((mes, 1))
dmenos = -np.ones((mes - 1, 1))

Difert = np.diag(np.ones(tri)) + np.diag(-np.ones(tri - 1), -1)
dmast = np.ones((tri, 1))
dmenost = -np.ones((tri - 1, 1))

W = np.zeros((tri * (ser - 1) + mes, 1))
n_colt = 1
n_colm = 1

# Carga de datos desde Excel
data = pd.read_excel(r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx", sheet_name="Nominal_a")
data.columns = data.columns.str.strip()

for category in ["CPR", "CPU", "INV", "EXP", "IMP", "AGRI", "INDU", "CST", "IMPU", "SER"]:
    if f"AN_{category}" not in data.columns:
        print(f"Advertencia: la columna AN_{category} no existe en el archivo.")
        continue

    p = data[f"AN_{category}"].values.reshape(-1, 1)

    if len(p) < tri:
        p = np.resize(p, (tri, 1))  # Asegurar tamaño correcto

    WT = np.kron(p[:tri], V4)  # Expandir con Kronecker
    WM = np.kron(p[:mes], v12)  # Expandir con Kronecker

    WDT = np.diag(WT[:tri, 0])  # Convertir en matriz diagonal
    WDM = np.diag(WM[:mes, 0])  # Convertir en matriz diagonal

    # **Matplace equivalente para HTW**
    row_start = 0
    row_end = min(WDT.shape[0], HTW.shape[0])
    col_start = n_colt - 1  # Convertir índice de EViews a Python (1-index → 0-index)
    col_end = col_start + tri

    if col_end <= HTW.shape[1]:  # Evitar desbordamiento de columnas
        HTW[row_start:row_end, col_start:col_end] = WDT[:row_end, :]
    else:
        print(f"Error en {category}: intento de asignar columnas fuera del rango en HTW")

    # **Matplace equivalente para HMW**
    row_start = 0
    row_end = min(WDM.shape[0], HMW.shape[0])
    col_start = n_colm - 1  # Convertir índice de EViews a Python
    col_end = col_start + mes

    if col_end <= HMW.shape[1]:  # Evitar desbordamiento de columnas
        HMW[row_start:row_end, col_start:col_end] = WDM[:row_end, :]
    else:
        print(f"Error en {category}: intento de asignar columnas fuera del rango en HMW")

    # Incrementar los contadores de columnas
    n_colt += tri
    n_colm += mes  # Solo avanza 228 en cada iteración

# Mostrar dimensiones finales
# print(f"HTW.shape: {HTW.shape}, HMW.shape: {HMW.shape}")

# **Insertar Bmagre en HMW**
row_start = mes # mes + 1 en indexado Python (0-based)
col_start = 0

row_end = row_start + Bmagre.shape[0]
col_end = col_start + Bmagre.shape[1]

# Verificar que Bmagre cabe en HMW en la posición especificada
if row_end <= HMW.shape[0] and col_end <= HMW.shape[1]:
    HMW[row_start:row_end, col_start:col_end] = Bmagre
else:
    print("Error: Bmagre no cabe en HMW en la posición especificada.")

# Mostrar dimensiones finales
#print(f"HTW.shape: {HTW.shape}, HMW.shape: {HMW.shape}")

# Carga de datos desde chowllin
data = pd.read_csv(r'C:\Users\lucas\High-Frequency_GDP_Forecasting\src\UNION\Ymens_final.csv', delimiter=";")
# data = pd.read_csv(r"C:\Users\lucas\Downloads\MAF_ESP\smoothed.csv", delimiter=",")  # Asegurar que "Periodo" es el índice
#data = data[data["_date_"].astype(str).isin([periodos[key] for key in ["201", "203"]])]
data['_date_'] = pd.to_datetime(data['_date_'])
data.set_index('_date_', inplace=True)
data = data.loc['2007-01-01':'2025-12-01']

# Seleccionar solo las 4 columnas deseadas
# = ["M_AGRI", "M_INDU", "M_CST", "M_IMPU", "M_SER"]  # Cambia por los nombres reales
#data = data[columnas_deseadas]

data.columns = data.columns.str.strip()

# Aplicación de HP Filter para suavizado
def apply_hp_filter(series, lamb=2):
    cycle, trend = hpfilter(series, lamb=lamb)
    return trend

for category in ["CPR", "CPU", "INV", "EXPOR", "IMP", 'AGRICUL', 'INDUSTRIA', 'CONSTRU', 'IMPUESTOS', 'SERVI']:
    if category in data.columns:
        data[f'{category}H'] = apply_hp_filter(data[category])
        #data.rename(columns={category: f"{category}H"}, inplace=True)  # Renombra la columna
    else:
        print(f"Advertencia: la columna {category} no existe en el archivo.")

#Cargar datos antes de bucle
# categorias_h = ["M_AGRIH", "M_INDUH", "M_CSTH", "M_IMPUH", "M_SERH"]
# categorias = ["M_AGRI", "M_INDU", "M_CST", "M_IMPU", "M_SER"]
# # Creación de estructuras de datos
# M_Y_ini = data[["M_CPRH", "M_CPUH", "M_INVH", "M_EXPH", "M_IMPH"]].copy()
# M_Y_inir = data[["M_CPR", "M_CPU", "M_INV", "M_EXP", "M_IMP"]].copy()
# # Generar las nuevas variables
# for categoria in categorias_h:
#     M_Y_ini[f'I_{categoria}'] = -data[categoria]
#     #M_Y_inir[f'I_{categoria}'] = -data[categoria]
#
# for categoria in categorias:
#     M_Y_inir[f'I_{categoria}'] = -data[categoria]

# Definir listas de categorías
categorias_h = ["AGRICULH", "INDUSTRIAH", "CONSTRUH", "IMPUESTOSH", "SERVIH"]
categorias = ["AGRICUL", "INDUSTRIA", "CONSTRU", "IMPUESTOS", "SERVI"]

# Creación de estructuras de datos
M_Y_ini = data[["CPRH", "CPUH", "INVH", "EXPORH", "IMPH"]].copy()
M_Y_inir = data[["CPR", "CPU", "INV", "EXPOR", "IMP"]].copy()

# Generar nuevas variables en un solo bucle
for cat_h, cat in zip(categorias_h, categorias):
    M_Y_ini[f'I_{cat_h}'] = -data[cat_h]
    M_Y_inir[f'I_{cat}'] = -data[cat]


#M_Y_inir = np.array(M_Y_inir)
T_Y_ini = np.dot(Bagre,M_Y_inir)
#T_Y_ini = Bagre @ M_Y_inir
#T_Y_ini = M_Y_inir * Bagre
T_Y_ini = pd.DataFrame(T_Y_ini)

# Convertir en un vector columna (similar a @vec en EViews)
VT_Y_ini = T_Y_ini.T.values.flatten().reshape(-1, 1)

# Extraer la submatriz (desde fila 0 hasta num_filas, columna 0)
#dem_T_ini = VT_Y_ini.iloc[0:tri * ser, :]  # En pandas

dem_T_ini = VT_Y_ini[:(tri * ser)//2, :]  # Si vT_Y_ini es un array de NumPy

# Calcular índices
start_row = (tri * ser) // 2  # En Python indexamos desde 0, por eso no sumamos 1
num_filas = (tri * ser) - start_row  # Asegurar que no excedemos el tamaño

# Si usas NumPy:
ofe_T_ini = -VT_Y_ini[start_row : start_row + num_filas, :]

# Ajuste de PIB
Difert = np.diag(dmast.flatten()) + np.diag(dmenost.flatten(), -1)
qt = Difert.T @ Difert
omegat = np.kron(np.eye(5), np.linalg.inv(qt))
# Cálculo de PIB con oferta y demanda
# Extraer HTW_d (submatriz de las primeras !tri filas y !tri*ser/2 columnas)
HTW_d = HTW[:tri, : (tri * ser) // 2]

# Extraer HTW_o (submatriz de las siguientes filas y columnas)
HTW_o = HTW[:tri, (tri * ser) // 2 : tri * ser]

PIB_dem_ini = HTW_d @ dem_T_ini
PIB_ofe_ini = HTW_o @ ofe_T_ini

PIB_T = p_dem * PIB_dem_ini + p_ofe * PIB_ofe_ini

pib = pd.read_excel(r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx", sheet_name="PIB", header=1)
pib = pib.rename(columns={"Unnamed: 0": "_date_"})
pib.set_index('_date_', inplace=True)
pib = pib['PIB']

pib = pd.DataFrame(pib)

# Calcular el crecimiento interanual (comparando con el trimestre del año pasado)
vtpib = ((pib["PIB"] - pib["PIB"].shift(4)) / pib["PIB"].shift(4)) * 100
vtpib = vtpib.loc['2007Q1':'2024Q3']

vtpib = vtpib.to_numpy().reshape(-1, 1)

PIB_T[:len(vtpib), 0] = vtpib[:, 0]  # Asignar vtpib a las primeras len(vtpib) filas de PIB_T

# Calcular la transpuesta de HTW_o y HTW_d
HTW_o_T = HTW_o.T
HTW_d_T = HTW_d.T

# Calcular la matriz inversa
try:
    inverse_term_o = np.linalg.inv(HTW_o @ omegat @ HTW_o_T)
    inverse_term_d = np.linalg.inv(HTW_d @ omegat @ HTW_d_T)
except np.linalg.LinAlgError:
    raise ValueError("La matriz no es invertible")

# Realizar las operaciones de la fórmula
DEM_T = dem_T_ini + omegat @ HTW_d_T @ inverse_term_d @ (PIB_T - (HTW_d @ dem_T_ini))
OFE_T = ofe_T_ini + omegat @ HTW_o_T @ inverse_term_o @ (PIB_T - (HTW_o @ ofe_T_ini))

# Extraer (submatriz de las siguientes filas y columnas)
dem_T_ini = VT_Y_ini[:(tri * ser)//2, :]  # Si vT_Y_ini es un array de NumPy

# Calcular índices
start_row = (tri * ser) // 2  # En Python indexamos desde 0, por eso no sumamos 1
num_filas = (tri * ser) - start_row  # Asegurar que no excedemos el tamaño

# Si usas NumPy:
ofe_T_ini = -VT_Y_ini[start_row : start_row + num_filas, :]

VT_Y = np.zeros((tri * ser, 1))
# VT_Y = VT_Y.to_numpy
VT_Y[:len(dem_T_ini), 0] = dem_T_ini[:, 0]  # Colocar el valor en la posición (1,1) -> índice 0 en Python

# Calcular la posición que corresponde al valor de tri * ser
index = int((tri * ser)/2) # En Python, el índice comienza en 0, por eso sumamos 1

# Sobrescribir los valores de VT_Y en la fila correspondiente al valor de tri * ser
VT_Y[index:index + len(ofe_T_ini), 0] = -ofe_T_ini[:, 0]

# Convertir en un vector columna (similar a @vec en EViews)
VM_Y_ini = M_Y_inir.T.values.flatten().reshape(-1, 1)

# Ajuste de PIB
Difer = np.diag(dmas.flatten()) + np.diag(dmenos.flatten(), -1)
q = Difer.T @ Difer
omega = np.kron(np.eye(10), np.linalg.inv(q))

# Según el formato original, parece que se empieza desde la fila 1, columna 1.
start_row = 1  # Fila de inicio para la extracción (basado en 1)
start_col = 1  # Columna de inicio (basado en 1)
end_row = int(tri * (ser - 1))  # El número de filas hasta donde se hace la extracción
end_col = 1  # Columna 1, de acuerdo a la fórmula

# Convertimos las coordenadas para que sean 0-basadas (en Python los índices empiezan desde 0)
start_row -= 1
start_col -= 1
end_row -= 1

# Extraer la submatriz de 'vt_Y' usando slicing
sub_vt_Y = VT_Y[start_row:end_row+1, start_col:end_col+1]


# Parámetros de ubicación
target_row_start = mes  # fila donde empieza (0-based)
target_row_end = target_row_start + sub_vt_Y.shape[0]
target_col = 0  # primera columna
# Insertar los datos
W[target_row_start:target_row_end, target_col] = sub_vt_Y.ravel()  # Asegura que sea vector plano

# Calcular la transpuesta de HMW
HMW_T = HMW.T

# Calcular la matriz inversa
try:
    inverse_term_ = np.linalg.inv(HMW @ omega @ HMW_T)
except np.linalg.LinAlgError:
    raise ValueError("La matriz no es invertible")

# Realizar las operaciones de la fórmula
VM_Y = VM_Y_ini + omega @ HMW_T @ inverse_term_ @ (W - (HMW @ VM_Y_ini))

# Crear matriz de salida
M_Y = np.zeros((mes, 10))

# Bucle para w de 1 a 10
for w in range(1, 11):  # 1 to 10 inclusive
    row_start = (w - 1) * (mes - 1) + w - 1  # índice base 0
    row_end = ((w - 1) * mes + mes)  # sin -1 porque Python slicing es exclusivo

    # Extraer submatriz de Vm_Y (columna 0 por defecto)
    sub = VM_Y[row_start:row_end, 0]

    # Verificamos que tenga longitud igual a 'mes' para insertar
    if sub.shape[0] == mes:
        M_Y[:, w - 1] = sub
    else:
        print(f"Advertencia: tamaño inesperado en w={w}, se esperaba {mes}, se obtuvo {sub.shape[0]}")

# DEMANDA Y OFERTA
demanda = VM_Y[0 : 5 * mes, :]
oferta = -VM_Y[5 * mes : , :]

# PESOS
pesodem = HMW[0 : mes, 0 : 5 * mes]
pesoOFE = HMW[0 : mes, 5 * mes : mes * (5 + ser)]

# PIBS
PIBDEM = pesodem @ demanda
PIBOFE = pesoOFE @ oferta

# Resultado con BAGRE
pib_t_INI = Bagre @ PIBDEM
PIB_T = Bagre @ PIBDEM
PIB_T[:len(vtpib), 0] = vtpib[:, 0]

DIF_PIB_T = PIB_T - pib_t_INI

tam = mes + tri
BFL = np.zeros((tam, tam))

# Colocar submatrices/vectores
BFL[0:q.shape[0], 0:q.shape[1]] = q
BFL[mes:, 0:Bagre.shape[1]] = Bagre
Bagre_T = Bagre.T
BFL[0:Bagre_T.shape[0], mes:] = Bagre.T
BFL[0, 0] = 1
# Vector ceroerr
ceroerr = np.zeros((tam, 1))
ceroerr[mes:, 0:] = DIF_PIB_T # Si dif_pib_t es escalar

# Resolver sistema lineal
suavizado = np.linalg.inv(BFL) @ ceroerr

# PIB_fin = PIBDEM + suavizado[0:mes, 0]
# Asegúrate de que PIB_fin sea un vector columna
# PIB_fin
# Sumar la primera columna de PIBDEM con la primera columna de suavizado
PIB_fin = pd.DataFrame(PIBDEM[:, 0] + suavizado[:mes, 0], columns=["PIB_fin"])
PIB_fin.index = pd.date_range("2007-01-01", periods=mes, freq="ME")

Finales = M_Y
Finales = pd.DataFrame(Finales)
Finales.columns = ["MF_CPR", "MF_CPU", "MF_INV", "MF_EXP", "MF_IMP", "I_MF_AGRI", "I_MF_INDU", "I_MF_CST", "I_MF_IMPU", "I_MF_SER"]
Finales.index = pd.date_range("2007-01-01", periods=mes, freq="ME")

# Seleccionar las columnas que comienzan con "I_MF"
cols_to_negate = Finales.filter(like="I_MF").columns

# Multiplicar dichas columnas por -1 y renombrarlas
Finales.rename(columns={col: col.replace("I_MF", "MF") for col in cols_to_negate}, inplace=True)
Finales[[col.replace("I_MF", "MF") for col in cols_to_negate]] *= -1

# Mostrar el DataFrame actualizado
Prediccion = Finales
# Agregar la columna PIB_fin al DataFrame Prediccion
Prediccion["PIB_fin"] = PIB_fin["PIB_fin"]

# Mover la columna PIB_fin al principio del DataFrame
Prediccion = Prediccion[['PIB_fin'] + [col for col in Prediccion.columns if col != 'PIB_fin']]

# Format as "YYYY-M"
Prediccion.index = Prediccion.index.to_series().dt.strftime("%YM%#m")
Prediccion.index.name = "OBS"

# Guardar el DataFrame en un archivo Excel
Prediccion.to_excel("prediccion1.xlsx", index=True, sheet_name="PREDICCION", float_format="%.6f")

# Rutas de los archivos
archivo_origen = "prediccion1.xlsx"
archivo_destino = r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx"
hoja_objetivo = "PREDICCION"

# Cargar libros
wb_origen = load_workbook(archivo_origen)
wb_destino = load_workbook(archivo_destino)

# Eliminar la hoja 'PREDICCION' en destino si ya existe
if hoja_objetivo in wb_destino.sheetnames:
    std = wb_destino[hoja_objetivo]
    wb_destino.remove(std)

# Copiar hoja desde origen (openpyxl no permite copiar directamente entre libros, así que recreamos)
ws_origen = wb_origen[hoja_objetivo]
ws_nueva = wb_destino.create_sheet(title=hoja_objetivo)

# Copiar celda a celda (contenido y formato básico)
for row in ws_origen.iter_rows():
    for cell in row:
        ws_nueva[cell.coordinate].value = cell.value

# Guardar el archivo destino con la hoja reemplazada
wb_destino.save(archivo_destino)
wb_origen.close()
wb_destino.close()
