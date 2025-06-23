import pandas as pd
import numpy as np
import statsmodels.api as sm
import warnings
import win32com.client
import time
import re
from openpyxl import load_workbook


config = {
    'CP18': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CP16': {
        'order': (1, 0, 1),
        'seasonal_order': (2, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CPU1': {
        'order': (1, 0, [0,1,1]),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CPU2': {
        'order': (1, 0, 0),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': False,
        'use_log_diff': True,
        # 'exog_extra': ['f20202311'],
    },
    'INV9': {
        'order': (2, 0, 0),
        'use_arima': True,
        'use_log_diff': False,
    },
    'EXP4': {
        'order': (1, 0, [0,1,0,0,0]),
        'seasonal_order': (3, 0, 0, 12),
        'use_arima': False,
        'use_log_diff': True,
        # 'exog_extra': ['fvirus','f202004'],
    },
    'EXP5': {
        'order': (3, 0, [0,0,0,0,0,0,0,0,0,0,0,1]),
        'use_arima': False,
        'use_log_diff': True,
    },
    'EXP6': {
        'order': (1, 0, 1),
        'seasonal_order': (3, 0, 0, 12),
        'use_arima': False,
        'use_log_diff': True,
        # 'exog_extra': ['f202004'],
    },
    'IND7': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
    },
    'IND8': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
    },
    'IND9': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CON9': {
        'order': (1, 0, [0,0,1]),
        'seasonal_order': (3, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CON10': {
        'order': (3, 0, 0),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CON11': {
        'order': (3, 0, 0),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'CON14': {
        'order': ([1,0,1], 0, 0),
        'seasonal_order': (2, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
    },
    'SER17': {
        'order': ([1,0,1], 0, 1),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
        # 'exog_extra': ['fvirus', 'f202003'],
    },
    'SER19': {
        'order': (2, 0, 0),
        'seasonal_order': (1, 0, 0, 12),
        'use_arima': True,
        'use_log_diff': False,
        # 'exog_extra': ['fvirus', 'f202003'],
    },
    'SER20': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
        # 'exog_extra': ['f202004', 'f202003'],
    },
    'SER21': {
        'order': (1, 0, 1),
        'use_arima': True,
        'use_log_diff': False,
        # 'exog_extra': ['f202003'],
    },
}

# Ignorar todos los warnings
warnings.filterwarnings("ignore")
from statsmodels.tsa.statespace.sarimax import SARIMAX

# Path to Excel file
ruta_excel = r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx"

# Launch Excel and open the file
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

# Open and update all links (UpdateLinks=3 updates everything)
wb = excel.Workbooks.Open(Filename=ruta_excel, UpdateLinks=3)

# Optional: Wait to ensure updates complete (useful for async links)
time.sleep(2)

# Save with updated values
wb.Save()
wb.Close()
excel.Quit()

# Carga de datos desde Excel
endogenas = pd.read_excel(r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx", sheet_name="LECTURA(PY)", index_col=0)
endogenas.columns = endogenas.columns.str.strip()

exogenas = pd.read_csv(r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\UNION\FactoresPCA\Factores_S_mensuales.csv", index_col=0, parse_dates=True)

def obtener_fechas_validas_por_columna(df,col):
    fechas_por_columna = {}
    # Filtra las fechas donde hay datos no nulos
    fechas_validas = df[col].dropna().index
    if len(fechas_validas) > 0:
        fechas_por_columna[col] = {
            'fecha_inicio': fechas_validas[0],
            'fecha_fin': fechas_validas[-1]
        }
    else:
        fechas_por_columna[col] = {
            'fecha_inicio': None,
            'fecha_fin': None
        }

    return fechas_por_columna

def interseccion_fechas(results_exo, results_endo):
    results = results_exo | results_endo
    fecha_inicio_comun = max(v['fecha_inicio'] for v in results.values())
    fecha_fin_comun = min(v['fecha_fin'] for v in results.values())
    return fecha_inicio_comun, fecha_fin_comun

def match_columns_by_prefix(df1, df2):
    matched = {}

    # Extract prefixes from df1 (e.g. 'CP1' -> 'CP')
    prefixes = set(re.sub(r'\d+', '', col) for col in df1.columns)

    # Iterate over each prefix
    for prefix in prefixes:
        # Build regex pattern: look for columns like FCP1F, FCP2F, etc.
        pattern = re.compile(rf'^F{prefix}\d*F?$')
        # Filter df2 columns matching the pattern
        matched_cols = [col for col in df2.columns if pattern.match(col)]
        matched[prefix] = matched_cols

    return matched

result = match_columns_by_prefix(endogenas, exogenas)

def obtener_intersecciones_de_fechas(df1, df2):
    resultados = {}

    # Obtener los prefijos del primer DataFrame
    prefijos = set(re.sub(r'\d+', '', col) for col in df1.columns)

    for prefijo in prefijos:
        # Seleccionar columnas de df2 que correspondan al prefijo (tipo FCP1F, FCP2F, etc.)
        pattern = re.compile(rf'^F{prefijo}\d*F?$')
        columnas_match = [col for col in df2.columns if pattern.match(col)]

        # Si no hay columnas asociadas, omitir
        if not columnas_match:
            continue

        # Obtener intersección de fechas comunes entre df1 y df2 (según índice)
        fechas_df1 = df1[[col for col in df1.columns if col.startswith(prefijo)]].dropna().index
        fechas_df2 = df2[columnas_match].dropna(how='all').index
        fechas_comunes = fechas_df1.intersection(fechas_df2)

        # Si no hay fechas comunes, saltar
        if fechas_comunes.empty:
            continue

        # Guardar fechas mínimas y máximas
        resultados[prefijo] = {
            "fecha_inicio": fechas_comunes.min(),
            "fecha_fin": fechas_comunes.max()
        }

    return resultados

results = obtener_intersecciones_de_fechas(endogenas, exogenas)

def safe_to_float(x):
    try:
        x_str = str(x).replace(',', '.').strip()
        if x_str in ['', '-', 'nan', 'None']:
            return np.nan
        return float(x_str)
    except:
        return np.nan

def forecast_log_diff_sarimax(df1, df2, fecha_forecast_fin, config=None):
    resultados_forecast = {}

    # Conversión segura de todos los valores a float
    df1 = df1.applymap(safe_to_float)
    df2 = df2.applymap(safe_to_float)

    for col in df1.columns:
        print(f"\nProcesando: {col}")

        # Extraer prefijo (ej. CP1 → CP)
        prefix_match = re.match(r'([A-Z]+)', col)
        if not prefix_match:
            print(f"Prefijo no reconocido para: {col}")
            continue
        prefijo = prefix_match.group(1)

        # Configuración personalizada si existe
        col_config = config.get(col, {}) if config else {}
        order = col_config.get('order', (1, 0, 1))
        seasonal_order = col_config.get('seasonal_order', (2, 0, 0, 12))
        exog_extra = col_config.get('exog_extra', [])
        use_arima = col_config.get('use_arima', False)
        use_log_diff = col_config.get('use_log_diff', True)

        # Columnas exógenas base + extras
        col_exog = [c for c in df2.columns if re.match(rf'^F{prefijo}\d*F$', c)]
        col_exog += [e for e in exog_extra if e in df2.columns]
        if not col_exog:
            print(f"Sin columnas exógenas para {col}")
            continue

        # Intersección de fechas
        fechas_endog = df1[col].dropna().index
        fechas_exog = df2[col_exog].dropna(how='all').index
        fechas_comunes = fechas_endog.intersection(fechas_exog)
        if fechas_comunes.empty:
            print(f"Sin fechas comunes para {col}")
            continue

        # Recorte de series
        fecha_inicio = fechas_comunes.min()
        fecha_fin = fechas_comunes.max()
        endog = df1.loc[fecha_inicio:fecha_fin, col]
        exog = df2.loc[fecha_inicio:fecha_fin, col_exog]

        # Intentar aplicar log-diferencia anual (solo si es posible)
        endog_shifted = endog.shift(12)
        valid_idx = (endog > 0) & (endog_shifted > 0)
        if use_log_diff:
            y = np.log(endog[valid_idx]) - np.log(endog_shifted[valid_idx])
        else:
            print(f"Log no aplicable para {col}. Usando serie directa sin transformación.")
            print(endog)
            print(endog_shifted[valid_idx])
            y = endog

        if y.empty:
            print(f"Serie vacía tras transformación: {col}")
            continue

        # Alinear exógenas
        train = pd.DataFrame({'y': y})
        exog.index.name = None  # Asegura compatibilidad
        exog_train = exog.reindex(train.index)
        if exog_train.empty:
            print(f"{col}: exog_train está vacío, se salta.")
            continue

        # Ajuste de modelo
        try:
            if use_arima:
                model = sm.tsa.ARIMA(train['y'],
                                     exog=exog_train,
                                     order=order,
                                     seasonal_order=seasonal_order,
                                     enforce_stationarity=False,
                                     enforce_invertibility=False)
                results = model.fit()
            else:
                model = sm.tsa.SARIMAX(train['y'],
                                       exog=exog_train,
                                       order=order,
                                       seasonal_order=seasonal_order,
                                       enforce_stationarity=False,
                                       enforce_invertibility=False)
                results = model.fit(disp=False)
        except Exception as e:
            print(f"Error ajustando modelo para {col}: {e}")
            continue

        # Forecast
        forecast_start = train.index[-1] + pd.DateOffset(months=1)
        forecast_index = pd.date_range(start=forecast_start,
                                       end=fecha_forecast_fin,
                                       freq='MS')
        exog_forecast = df2.loc[forecast_index, col_exog]

        forecast = results.get_forecast(steps=len(forecast_index),
                                        exog=exog_forecast)
        forecast_mean = forecast.predicted_mean

        # Reconstrucción de la serie original
        valor_base = df1.loc[forecast_start - pd.DateOffset(months=12), col]
        reconstruido = valor_base * np.exp(forecast_mean.cumsum())
        reconstruido.name = col
        if use_log_diff:
            resultado = reconstruido
        else:
            resultado = forecast_mean
        resultados_forecast[col] = {
            'resultado': resultado
        }
        print(resultados_forecast[col])
    return resultados_forecast

resultado = forecast_log_diff_sarimax(endogenas, exogenas, '2025-12-01',config)

# Crear el DataFrame con los forecasts (asegurando columnas planas)
df_forecast = pd.concat({k: v['resultado'] for k, v in resultado.items()}, axis=1)
df_forecast.columns.name = None

# Unificar columnas sin eliminar ninguna
columnas_finales = endogenas.columns.union(df_forecast.columns)
endogenas = endogenas.reindex(columns=columnas_finales)
df_forecast = df_forecast.reindex(columns=columnas_finales)

# Rellenar solo las fechas ya presentes en endogenas (como 2025-xx) que tienen NaN
# Esto mantiene el mismo índice sin duplicarlo
endogenas_actualizado = endogenas.copy()
endogenas_actualizado.update(df_forecast)

# Eliminar la columna no deseada
endogenas_actualizado = endogenas_actualizado.drop(columns=['PIBUEM'], errors='ignore')

# Format as "YYYY-M"
endogenas_actualizado.index = endogenas_actualizado.index.to_series().dt.strftime("%YM%#m")
endogenas_actualizado.index.name = "OBS"

# Guardar
output_path = r'C:\Users\lucas\High-Frequency_GDP_Forecasting\src\UNION\PREDINDI.xlsx'
endogenas_actualizado.to_excel(output_path, sheet_name='PREDINDI', index=True, float_format="%.6f")

print(f"✅ Forecast rellenado en fechas existentes. Archivo guardado en:\n{output_path}")

# Rutas de los archivos
archivo_origen = "PREDINDI.xlsx"
archivo_destino = r"C:\Users\lucas\High-Frequency_GDP_Forecasting\src\data\AFESP.xlsx"
hoja_objetivo = "PREDINDI"

# Cargar libros
wb_origen = load_workbook(archivo_origen)
wb_destino = load_workbook(archivo_destino)

# Eliminar la hoja 'PREDICCION' en destino si ya existe
if hoja_objetivo in wb_destino.sheetnames:
    std = wb_destino[hoja_objetivo]
    wb_destino.remove(std)

# Copiar hoja desde origen (openpyxl no permite copiar directamente entre libros, así que recreamos)
ws_origen = wb_origen[hoja_objetivo]
ws_nueva = wb_destino.create_sheet(title=hoja_objetivo)

# Copiar celda a celda (contenido y formato básico)
for row in ws_origen.iter_rows():
    for cell in row:
        ws_nueva[cell.coordinate].value = cell.value

# Guardar el archivo destino con la hoja reemplazada
wb_destino.save(archivo_destino)
wb_origen.close()
wb_destino.close()